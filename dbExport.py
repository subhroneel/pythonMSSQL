# ! /usr/bin/env python3
import os
from tkinter.ttk import *

import cx_Oracle
from pyodbc import Cursor
# import psycopg2
import pyodbc
import openpyxl.workbook
from openpyxl import Workbook
from tkinter.messagebox import showinfo
from tkinter.simpledialog import askinteger
from tkinter import *

from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.worksheet import Worksheet


# from pyodbc import Connection, IntegrityError


# import cxOracle

# from pyodbc import *


class dbExport:
    def __init__(self, root):
        self.root = root
        self.root.title('Database Exporter')
        self.root.geometry('850x300')
        self.totrec: int = 0
        # self.conn: Connection = None
        # self.tableList: list

        srvFrame = LabelFrame(self.root, text="List of Database Servers", font=("Arial", 12, "bold"), bg="#4d65a3",
                              fg="white",
                              bd=5, relief=GROOVE)
        srvFrame.place(x=0, y=0, width=400, height=50)
        self.Combo = Combobox(srvFrame, width=27, textvariable=StringVar(), values=['MSSQL Server 2022', 'Oracle 21c', 'MySQL Server', 'PostgreSQL Server'])
        self.Combo.set("")
        self.Combo.bind('<<ComboboxSelected>>', self.getAllSchemas)
        self.Combo.pack(side=TOP, fill=Y)

        schemaFrame = LabelFrame(self.root, text="List of Schemas", font=("Arial", 12, "bold"), bg="#0f6751",
                                 fg="white",
                                 bd=5, relief=GROOVE)
        schemaFrame.place(x=0, y=55, width=400, height=60)
        self.schemaCombo = Combobox(schemaFrame, width=27, textvariable=StringVar(),
                                    values=[''])
        self.schemaCombo.set("")
        self.schemaCombo.bind('<<ComboboxSelected>>', self.getAllTables)
        self.schemaCombo.pack(side=TOP, fill=Y)

        buttonExcelFrame = LabelFrame(self.root, text="Export to Excel", font=("arial", 12, "bold"), bg="#df7d44",
                                      fg="white",
                                      bd=5, relief=GROOVE)
        buttonExcelFrame.place(x=0, y=120, width=200, height=60)
        self.exportExcelBtn = Button(buttonExcelFrame, text="Export", command=lambda: self.exportToExcel(), width=14,
                                     height=2,
                                     font=("arial", 14, "bold"),
                                     fg="#9ed6e6", bg="#4c4f55")
        self.exportExcelBtn.grid(row=0, column=0, padx=0, pady=5)
        self.exportExcelBtn.pack()
        self.exportExcelBtn["state"] = DISABLED

        buttonCSVFrame = LabelFrame(self.root, text="Export to CSV", font=("arial", 12, "bold"), bg="#df7d44",
                                    fg="white",
                                    bd=5, relief=GROOVE)
        buttonCSVFrame.place(x=205, y=120, width=200, height=60)
        self.exportCSVBtn = Button(buttonCSVFrame, text="Export", command=lambda: self.exportToCSV(), width=14,
                                   height=2,
                                   font=("arial", 14, "bold"),
                                   fg="#9ed6e6", bg="#4c4f55")
        self.exportCSVBtn.grid(row=0, column=0, padx=0, pady=5)
        self.exportCSVBtn.pack()
        self.exportCSVBtn["state"] = DISABLED

        pbframe = LabelFrame(self.root, text="", font=("arial", 12, "bold"), bg="#8F00FF", fg="white",
                             bd=5, relief=GROOVE)
        pbframe.place(x=0, y=200, width=400, height=40)
        self.pb = Progressbar(pbframe, orient='horizontal', mode='determinate', length=280)
        # self.pb.grid(column=0, row=0, columnspan=2, padx=20, pady=40)
        self.pb.pack()
        self.pb['value'] = 0
        tblListframe = LabelFrame(self.root, text="List of Tables", font=("arial", 12, "bold"), bg="#8F00FF",
                                  fg="white",
                                  bd=5, relief=GROOVE)
        tblListframe.place(x=450, y=0, width=400, height=200)
        scroll_y = Scrollbar(tblListframe, orient=VERTICAL)
        self.LB = Listbox(tblListframe, yscrollcommand=scroll_y.set, selectbackground="#8d8df6", selectmode=EXTENDED,
                          font=("arial", 12, "bold"), bg="#c9f56f", fg="navyblue", bd=5, relief=GROOVE)
        scroll_y.config(command=self.LB.yview)
        scroll_y.pack(side=RIGHT, fill=Y)
        self.LB.bind('<<ListboxSelect>>', self.items_selected)
        self.LB.pack(fill=BOTH)
        self.LB['state'] = DISABLED

    def convertCursorRowToListForTable(self, result: list):
        row_to_list = []
        # result1 = [list(rows) for rows in result]
        if self.Combo.get() == "MSSQL Server 2022":
            for row in result:
                row_to_list.append(row.table_name)
        elif self.Combo.get() == "Oracle 21c":
            for row in result:
                row_to_list.append(row[0])
        elif self.Combo.get() == "MySQL Server":
            for row in result:
                row_to_list.append(row.table_name)
        elif self.Combo.get() == "PostgreSQL Server":
            for row in result:
                row_to_list.append(row.table_name)
        return row_to_list

    def convertCursorRowToListForSchema(self, result: list):
        row_to_list = []
        # result1 = [list(rows) for rows in result]
        if self.Combo.get() == "MSSQL Server 2022":
            for row in result:
                row_to_list.append(row.dbname)
        elif self.Combo.get() == "Oracle 21c":
            for row in result:
                row_to_list.append(row[0])
        elif self.Combo.get() == "MySQL Server":
            for row in result:
                row_to_list.append(row.dbname)
        elif self.Combo.get() == "PostgreSQL Server":
            for row in result:
                row_to_list.append(row.dbname)
        return row_to_list

    def convertCursorRowToListForColumn(self, result: list):
        row_to_list = []
        # result1 = [list(rows) for rows in result]
        if self.Combo.get() == "MSSQL Server 2022":
            for row in result:
                row_to_list.append(row.column_name)
        elif self.Combo.get() == "Oracle 21c":
            for row in result:
                row_to_list.append(row[0])
        elif self.Combo.get() == "MySQL Server":
            for row in result:
                row_to_list.append(row.column_name)
        elif self.Combo.get() == "PostgreSQL Server":
            for row in result:
                row_to_list.append(row.column_name)
        return row_to_list

    def convertCursorRowToListForDataType(self, result: list):
        row_to_list = []
        # result1 = [list(rows) for rows in result]
        if self.Combo.get() == "MSSQL Server 2022":
            for row in result:
                row_to_list.append(row.data_type)
        elif self.Combo.get() == "Oracle 21c":
            for row in result:
                row_to_list.append(row[0])
        elif self.Combo.get() == "MySQL Server":
            for row in result:
                row_to_list.append(row.data_type)
        elif self.Combo.get() == "PostgreSQL Server":
            for row in result:
                row_to_list.append(row.data_type)
        return row_to_list

    def getTableRecordSet(self):
        cursor: Cursor = self.conn.cursor()
        if self.Combo.get() == "MSSQL Server 2022":
            cursor.execute('USE ' + self.schemaName)
            cursor.execute('SELECT trim(upper(table_name)) as table_name FROM information_schema.tables where '
                           'table_catalog = ? ', self.schemaName)
        elif self.Combo.get() == "Oracle 21c":
            cursor.execute(
                'SELECT trim(upper(table_name)) as table_name FROM ALL_TABLES WHERE OWNER  = \'' + self.schemaName + '\'')
        elif self.Combo.get() == "MySQL Server":
            cursor.execute('SELECT trim(upper(table_name)) as table_name FROM information_schema.tables where '
                           'upper(table_schema) = ? ', self.schemaName)
        elif self.Combo.get() == "PostgreSQL Server":
            cursor.execute('SELECT trim(upper(table_name)) as table_name FROM information_schema.tables where '
                           'table_type = \'BASE TABLE\' and upper(table_schema) = \'PUBLIC\' and upper(table_catalog) '
                           '= ? ', self.schemaName)
        result = cursor.fetchall()
        cursor.close()
        return result

    def getColumnnNameForTable(self, tbl_name: str):
        cursor: Cursor = self.conn.cursor()
        if self.Combo.get() == "MSSQL Server 2022":
            cursor.execute('SELECT upper(column_name) as column_name FROM information_schema.columns '
                           'where upper(table_catalog) = ? and trim(upper(table_name)) = ? order by ordinal_position',
                           self.schemaName, tbl_name)
        elif self.Combo.get() == "Oracle 21c":
            cursor.execute('SELECT upper(column_name) as column_name FROM all_tab_columns '
                           'where upper(owner) = \'' + self.schemaName + '\' and trim(upper(table_name)) = \'' + tbl_name + '\'')
        elif self.Combo.get() == "MySQL Server":
            cursor.execute('SELECT upper(column_name) as column_name FROM information_schema.columns '
                           'where upper(table_schema) = ? and trim(upper(table_name)) = ? order by ordinal_position',
                           self.schemaName, tbl_name)
        elif self.Combo.get() == "PostgreSQL Server":
            cursor.execute('SELECT upper(column_name) as column_name FROM information_schema.columns '
                           'where table_schema=\'public\' and upper(table_catalog) = ? and trim(upper(table_name)) = '
                           '? order by ordinal_position',
                           self.schemaName, tbl_name)
        result = cursor.fetchall()
        cursor.close()
        return result

    def getColumnnDataTypeForTable(self, tbl_name: str):
        cursor = self.conn.cursor()
        if self.Combo.get() == "MSSQL Server 2022":
            cursor.execute('SELECT upper(data_type) as data_type FROM information_schema.columns '
                           'where upper(table_catalog) = ? and trim(upper(table_name)) = ? order by ordinal_position',
                           self.schemaName, tbl_name)
        elif self.Combo.get() == "Oracle 21c":
            cursor.execute('SELECT upper(data_type) as data_type FROM all_tab_columns '
                           'where upper(owner) = \'' + self.schemaName + '\' and trim(upper(table_name)) = \'' + tbl_name + '\'')
        elif self.Combo.get() == "MySQL Server":
            cursor.execute('SELECT upper(data_type) as data_type FROM information_schema.columns '
                           'where upper(table_schema) = ? and trim(upper(table_name)) = ? order by ordinal_position',
                           self.schemaName, tbl_name)
        elif self.Combo.get() == "PostgreSQL Server":
            cursor.execute('SELECT upper(data_type) as data_type FROM information_schema.columns '
                           'where table_schema=\'public\' and upper(table_catalog) = ? and trim(upper(table_name)) = '
                           '? order by ordinal_position',
                           self.schemaName, tbl_name)
        result = cursor.fetchall()
        cursor.close()
        return result
    def getColumnRecordSet(self):
        table_type = 'BASE TABLE'
        allddl: str = ''
        cursor = self.conn.cursor()
        if self.Combo.get() == "MSSQL Server 2022":
            cursor.execute('SELECT upper(table_name) table_name,upper(column_name) column_name,'
                           'upper(data_type) data_type,upper(is_nullable) is_nullable,'
                           'character_maximum_length,numeric_precision,numeric_scale '
                           'FROM information_schema.columns '
                           'where trim(upper(table_name)) in (?) order by table_name, ordinal_position', self.tableList)
        elif self.Combo.get() == "Oracle 21c":
            cursor.execute('SELECT upper(table_name) table_name,upper(column_name) column_name,'
                           'upper(data_type) data_type,upper(nullable) is_nullable,'
                           'data_length,data_precision,data_scale '
                           'FROM user_tab_columns '
                           'where trim(upper(table_name)) in (?) order by table_name, column_id', self.tableList)
        result = cursor.fetchall()
        cursor.close()
        return result

    def getConnection(self, dataSrc: str):
        if self.Combo.get() == "MSSQL Server 2022":
            ConnectionString = (
                "DRIVER={ODBC Driver 18 for MSSQL Server};SERVER=127.0.0.1;UID=sa;PWD=P@ssw0rd"
                ";TrustServerCertificate=yes;")
            return pyodbc.connect(ConnectionString)
        elif self.Combo.get() == "Oracle 21c":
            return cx_Oracle.connect("system/S0m0$hr33@192.168.29.234")
            # return cx_Oracle.connect("system/S0m0shr33@192.168.29.234")
        elif self.Combo.get() == "MySQL Server":
            ConnectionString = (
                "DRIVER={MySQL ODBC 8.2 Unicode Driver}; SERVER=192.168.29.127;DATABASE=greendb; UID=neel; "
                "PASSWORD=P@ssw0rd;")
        elif self.Combo.get() == "PostgreSQL Server":
            ConnectionString = (
                "DRIVER=/usr/local/lib/psqlodbcw.so;SERVER=192.168.29.127;DATABASE=greendb;UID=postgres;PASSWORD=P"
                "@ssw0rd;")
        return pyodbc.connect(ConnectionString)

        # elif Combo.get() == "PostgresSQL": conn = psycopg2.connect(host="localhost",port=5433,database="your_database",
        # user="your_username",password="your_password")

    def getTableRecords(self, table_name: str, columnNames: list):
        cursor = self.conn.cursor()
        colNames: str = ','.join(columnNames)
        if self.Combo.get() == "MSSQL Server 2022":
            cursor.execute('SELECT ' + colNames + ' FROM ' + self.schemaName + '..' + table_name)
        elif self.Combo.get() == "Oracle 21c":
            cursor.execute('SELECT ' + colNames + ' FROM ' + self.schemaName + '.' + table_name)
        elif self.Combo.get() == "MySQL Server":
            cursor.execute('SELECT ' + colNames + ' FROM ' + self.schemaName.lower() + '.' + table_name.lower())
        elif self.Combo.get() == "PostgreSQL Server":
            cursor.execute('SELECT * FROM ' + self.schemaName.lower() + '.public.' + table_name.lower())
        result = cursor.fetchall()
        cursor.close()
        return result

    def getTotTableRecordsCount(self):
        if self.tableList != "":
            cursor = self.conn.cursor()
            sqlStr: str = "select sum(cnt) as totcount from ("
            tbllist = self.tableList.split(",")
            if self.Combo.get() == "MSSQL Server 2022":
                for table_name in tbllist:
                    sqlStr += 'select count(*) as cnt from ' + self.schemaName + '..' + table_name + ' UNION '
                sqlStr += 'select 0) tbl'
            elif self.Combo.get() == "Oracle 21c":
                for table_name in tbllist:
                    sqlStr += 'select count(*) as cnt from ' + self.schemaName + '.' + table_name + ' UNION '
                sqlStr += 'select 0 from dual) tbl'
            elif self.Combo.get() == "MySQL Server":
                for table_name in tbllist:
                    sqlStr += 'select count(*) as cnt from ' + self.schemaName + '.' + table_name + ' UNION '
            elif self.Combo.get() == "PostgreSQL Server":
                for table_name in tbllist:
                    sqlStr += 'select count(*) as cnt from ' + self.schemaName + '.PUBLIC.' + table_name + ' UNION '
                sqlStr += 'select 0) tbl'
            cursor.execute(sqlStr)
            result = cursor.fetchall()
            cursor.close()
            for row in result:
                return row.totcount if self.Combo.get() == "MSSQL Server 2022" else row[0]

    # def addCheckBoxToListBox(self, frame: Frame, LB: Listbox, list_for_listbox: list):
    #     enable = []
    #     i = 1
    #     for item in list_for_listbox:
    #         enable.append("button " + str(i))
    #         i = i + 1
    #     for item in list_for_listbox:
    #         for y in enable:
    #             globals()["var{}{}".format(item, y)] = BooleanVar()
    #             globals()["checkbox{}{}".format(item, y)] = Checkbutton(frame, text=y,
    #                                                                     variable=globals()["var{}{}".format(item, y)])

    def addTableToListBox(self, result: list):
        self.LB['state'] = NORMAL
        self.LB.delete(0, END)
        for item in result:
            self.LB.insert(END, item)

    def items_selected(self, event):
        # get selected indices
        selected_indices = self.LB.curselection()
        w = event.widget
        # get selected items
        selected_langs = ",".join([self.LB.get(i) for i in selected_indices])
        self.tableList = selected_langs
        if len(self.tableList) > 0:
            self.exportExcelBtn['state'] = NORMAL
            self.exportCSVBtn['state'] = NORMAL
        else:
            self.exportExcelBtn['state'] = DISABLED
            self.exportCSVBtn['state'] = DISABLED

    def getExcelColumnIndex(self, columnIndex: int) -> str:
        if int((columnIndex - 1) / 26) > 0:
            s = chr(64 + int(columnIndex / 26))
        else:
            s = ''
        return s + chr(65 + ((columnIndex - 1) % 26))

    def resizeCells(self, ws: Worksheet):
        for col in ws.columns:
            SetLen = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                if len(str(cell.value)) > SetLen:
                    SetLen = len(str(cell.value))
            # Setting the column width
            ws.column_dimensions[column].width = SetLen + 5

    def exportToExcel(self):
        # result = getTableRecordSet(conn)
        if self.tableList != "":
            tbllist = self.tableList.split(",")
            wb = Workbook()
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            self.pb['maximum'] = self.getTotTableRecordsCount()
            self.pb.start()
            self.pb['value'] = 0
            self.pb.start()
            for table_name in tbllist:
                ws = wb.create_sheet(table_name)
                columnNames = self.convertCursorRowToListForColumn(self.getColumnnNameForTable(table_name))
                columnDataTypes = self.convertCursorRowToListForDataType(self.getColumnnDataTypeForTable(table_name))
                result1 = self.getTableRecords(table_name, columnNames)
                r = 1
                c = 1
                SetLen = 0
                for col in columnNames:
                    ws.cell(row=r, column=c).value = col
                    ws.cell(row=r, column=c).font = Font(size=12, bold=True)
                    ws.cell(row=r, column=c).border = thin_border
                    c += 1

                for row1 in result1:
                    r = r + 1
                    c = 1
                    for x in row1:
                        ws.cell(row=r, column=c).value = x
                        ws.cell(row=r, column=c).font = Font(size=9, bold=False)
                        ws.cell(row=r, column=c).border = thin_border
                        c += 1
                    self.pb['value'] += 1
                    self.root.update_idletasks()
                self.resizeCells(ws)
            self.pb.stop()
            wb.save(self.schemaName + ".xlsx")
            message = f'Export completed successfully for  {self.Combo.get()}!'
            self.exportExcelBtn['state'] = DISABLED
            self.exportCSVBtn['state'] = DISABLED

    def makeDirectory(self):
        directory = self.schemaName
        # Parent Directory path
        parent_dir = "./"
        # Path
        path = os.path.join(parent_dir, directory)
        try:
            os.mkdir(path)
        except OSError as error:
            print(error)
        return path

    def exportToCSV(self):
        # result = getTableRecordSet(conn)
        if self.tableList != "":
            tbllist = self.tableList.split(",")
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            self.pb['maximum'] = self.getTotTableRecordsCount()
            self.pb.start()
            self.pb['value'] = 0
            self.pb.start()
            self.path = self.makeDirectory()
            for table_name in tbllist:
                file1 = open(self.path + '/' + table_name + '.csv', 'w')
                columnNames = self.convertCursorRowToListForColumn(self.getColumnnNameForTable(table_name))
                columnDataTypes = self.convertCursorRowToListForDataType(self.getColumnnDataTypeForTable(table_name))
                result1 = self.getTableRecords(table_name, columnNames)
                r = 1
                c = 1
                SetLen = 0
                header: str = ''
                record: str = ''
                for col in columnNames:
                    header += (col + '~')
                file1.writelines(header[:-1] + '\n')
                for row1 in result1:
                    r = r + 1
                    c = 1
                    line: str = ''
                    for x in row1:
                        line += (str(x) + '~')
                        c += 1
                    file1.writelines(line[:-1] + '\n')
                    self.pb['value'] += 1
                    self.root.update_idletasks()
                file1.close()
            self.pb.stop()
            message = f'Export completed successfully for  {self.Combo.get()}!'
            self.exportExcelBtn['state'] = DISABLED
            self.exportCSVBtn['state'] = DISABLED

    def getAllTables(self, event):
        # self.conn = self.getConnection(self.Combo.get())
        self.schemaName = self.schemaCombo.get()
        result = self.convertCursorRowToListForTable(self.getTableRecordSet())
        self.addTableToListBox(result)
        self.LB.pack(expand=True, fill=BOTH, side=LEFT)

    def getAllSchemas(self, event):
        self.conn = self.getConnection(self.Combo.get())
        cursor: Cursor = self.conn.cursor()
        if self.Combo.get() == "MSSQL Server 2022":
            cursor.execute('SELECT distinct upper(name) as dbname FROM sys.databases')
        elif self.Combo.get() == "Oracle 21c":
            cursor.execute('SELECT distinct upper(USERNAME) as dbname FROM ALL_USERS WHERE ORACLE_MAINTAINED=\'N\'')
        elif self.Combo.get() == "MySQL Server":
            cursor.execute('SELECT distinct upper(SCHEMA_NAME) as dbname FROM INFORMATION_SCHEMA.SCHEMATA WHERE '
                           'schema_name not in (\'information_schema\',\'performance_schema\',\'sys\')')
        elif self.Combo.get() == "PostgreSQL Server":
            cursor.execute('SELECT distinct upper(CATALOG_NAME) as dbname FROM INFORMATION_SCHEMA.SCHEMATA WHERE '
                           'schema_name in (\'public\')')
        result = cursor.fetchall()
        self.schemaCombo['values'] = self.convertCursorRowToListForSchema(result)
        cursor.close()

root = Tk()
dbExport(root)
print(pyodbc.drivers())
root.mainloop()
