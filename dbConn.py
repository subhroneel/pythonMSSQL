# ! /usr/bin/env python3
import os
import tkinter as tk

import cx_Oracle
import openpyxl.workbook
from openpyxl import Workbook
from tkinter.messagebox import showinfo
from tkinter.simpledialog import askinteger
from tkinter import *
from tkinter import messagebox, ttk

from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side

top = Tk()
from pyodbc import Connection
# import psycopg2
import pyodbc


# import cxOracle

# from pyodbc import *
def convertCursorRowToListForTable(result: list):
    row_to_list = []
    # result1 = [list(rows) for rows in result]
    if Combo.get() == "MSSQL Server 2022":
        for row in result:
            row_to_list.append(row.table_name)
    elif Combo.get() == "Oracle 21c":
        for row in result:
            row_to_list.append(row[0])
    return row_to_list


def convertCursorRowToListForColumn(result: list):
    row_to_list = []
    # result1 = [list(rows) for rows in result]
    if Combo.get() == "MSSQL Server 2022":
        for row in result:
            row_to_list.append(row.column_name)
    elif Combo.get() == "Oracle 21c":
        for row in result:
            row_to_list.append(row[0])
    return row_to_list


def getTableRecordSet(conn: Connection):
    cursor = conn.cursor()
    if Combo.get() == "MSSQL Server 2022":
        cursor.execute('SELECT trim(upper(table_name)) as table_name FROM information_schema.tables')
    elif Combo.get() == "Oracle 21c":
        cursor.execute('SELECT trim(upper(table_name)) as table_name FROM USER_TABLES')
    result = cursor.fetchall()
    cursor.close()
    return result


def getColumnnNameForTable(conn: Connection, tbl_name: str):
    cursor = conn.cursor()
    if Combo.get() == "MSSQL Server 2022":
        cursor.execute('SELECT upper(column_name) as column_name FROM information_schema.columns '
                       'where trim(upper(table_name)) = ? order by ordinal_position', tbl_name)
    elif Combo.get() == "Oracle 21c":
        cursor.execute('SELECT upper(column_name) as column_name FROM user_tab_columns '
                       'where trim(upper(table_name)) = \'' + tbl_name + '\'')
    result = cursor.fetchall()
    cursor.close()
    return result


def getColumnRecordSet(conn: Connection):
    table_type = 'BASE TABLE'
    allddl: str = ''
    cursor = conn.cursor()
    if Combo.get() == "MSSQL Server 2022":
        cursor.execute('SELECT upper(table_name) table_name,upper(column_name) column_name,'
                       'upper(data_type) data_type,upper(is_nullable) is_nullable,'
                       'character_maximum_length,numeric_precision,numeric_scale '
                       'FROM information_schema.columns '
                       'where trim(upper(table_name)) in (?) order by table_name, ordinal_position', tableList)
    elif Combo.get() == "Oracle 21c":
        cursor.execute('SELECT upper(table_name) table_name,upper(column_name) column_name,'
                       'upper(data_type) data_type,upper(nullable) is_nullable,'
                       'data_length,data_precision,data_scale '
                       'FROM user_tab_columns '
                       'where trim(upper(table_name)) in (?) order by table_name, column_id', tableList)
    result = cursor.fetchall()
    cursor.close()
    return result


def getConnection(dataSrc: str):
    if Combo.get() == "MSSQL Server 2022":
        ConnectionString = (
            "DRIVER={ODBC Driver 18 for SQL Server};SERVER=127.0.0.1;DATABASE=greendb;UID=subhro;PWD=Subhr09!l"
            ";TrustServerCertificate=yes;")
        return pyodbc.connect(ConnectionString)
    elif Combo.get() == "Oracle 21c":
        return cx_Oracle.connect("c##greendb/greendb@192.168.29.234")
    # elif Combo.get() == "PostgresSQL": conn = psycopg2.connect(host="localhost",port=5433,database="your_database",
    # user="your_username",password="your_password")


def getTableRecords(conn: Connection, table_name: str, columnNames: list):
    cursor = conn.cursor()
    colNames: str = ','.join(columnNames)
    if Combo.get() == "MSSQL Server 2022":
        cursor.execute('SELECT ' + colNames + ' FROM GREENDB..' + table_name)
    elif Combo.get() == "Oracle 21c":
        cursor.execute('SELECT ' + colNames + ' FROM C##GREENDB.' + table_name)
    result = cursor.fetchall()
    cursor.close()
    return result


# conn = getConnection()
# printData(conn)
# def insertTableIntoListBox(result: list):
#     for row in result:
#         row_to_list.append(row.table_name)
#     return row_to_list

def addCheckBoxToListBox(frame: Frame, LB: Listbox, list_for_listbox: list):
    enable = []
    i = 1
    for item in list_for_listbox:
        enable.append("button " + str(i))
        i = i + 1
    for item in list_for_listbox:
        for y in enable:
            globals()["var{}{}".format(item, y)] = BooleanVar()
            globals()["checkbox{}{}".format(item, y)] = Checkbutton(frame, text=y,
                                                                    variable=globals()["var{}{}".format(item, y)])


def addTableToListBox(result: list):
    for item in result:
        LB.insert(END, item)


def items_selected(event):
    # get selected indices
    selected_indices = LB.curselection()
    w = event.widget
    # get selected items
    selected_langs = ",".join([LB.get(i) for i in selected_indices])
    # msg = f'You selected: {selected_langs}'
    # showinfo(title='Information', message=msg)
    global tableList
    tableList = selected_langs
    if len(tableList) > 0:
        btn.state(['!disabled'])
    else:
        btn.state(['disabled'])


def getExcelColumnIndex(columnIndex: int) -> str:
    if int((columnIndex - 1) / 26) > 0:
        s = chr(64 + int(columnIndex / 26))
    else:
        s = ''
    return s + chr(65 + ((columnIndex - 1) % 26))


def exportToExcel(result: list, conn: Connection):
    # result = getTableRecordSet(conn)
    tbllist = tableList.split(",")
    wb = Workbook()
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for table_name in tbllist:
        ws = wb.create_sheet(table_name)
        columnNames = convertCursorRowToListForColumn(getColumnnNameForTable(conn, table_name))
        result1 = getTableRecords(conn, table_name, columnNames)
        r = 1
        c = 1
        SetLen = 0
        for col in columnNames:
            ws.cell(row=r, column=c).value = col
            ws.cell(row=r, column=c).font = Font(size=12, bold=True)
            ws.cell(row=r, column=c).border = thin_border
            c = c + 1

        for row1 in result1:
            r = r + 1
            c = 1
            for x in row1:
                ws.cell(row=r, column=c).value = x
                ws.cell(row=r, column=c).font = Font(size=9, bold=False)
                ws.cell(row=r, column=c).border = thin_border
                c = c + 1
    for col in ws.columns:
        SetLen = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            if len(str(cell.value)) > SetLen:
                SetLen = len(str(cell.value))
        # Setting the column width
        ws.column_dimensions[column].width = SetLen + 5
    wb.save("greendborcl1.xlsx")

def ComboBoxSelect(event):
    conn = getConnection(Combo.get())
    result = convertCursorRowToListForTable(getTableRecordSet(conn))
    addTableToListBox(result)
    LB.pack(expand=True, fill=tk.BOTH, side=tk.LEFT)
    global btn
    btn = ttk.Button(top, text="Export to Excel", command=lambda: exportToExcel(result, conn))
    btn.pack(ipadx=5, ipady=5, expand=True)


def Apps():
    # def __init__(self, master):
    #     super().__init__()
    global top
    top = tk.Tk()
    top.geometry("400x250")

    frame = Frame(top)
    frame.pack()
    global Combo
    Combo = ttk.Combobox(frame, values=['MSSQL Server 2022', 'Oracle 21c'])
    #Combo.set("MSSQL Server 2022")
    Combo.set("")
    Combo.pack(padx=0, pady=5)
    Combo.bind('<<ComboboxSelected>>', ComboBoxSelect)
    global LB
    LB = Listbox(frame, height=10, selectmode=tk.EXTENDED)
    # LB.grid(row=1, column=1)
    # scrollbar = ttk.Scrollbar(
    #     top,
    #     orient=tk.VERTICAL,
    #     command=LB.yview)
    # LB['yscrollcommand'] = scrollbar.set
    # # addCheckBoxToListBox(frame, LB, result)
    # scrollbar.pack(side=tk.LEFT, expand=True, fill=tk.Y)
    LB.bind('<<ListboxSelect>>', items_selected)
    top.mainloop(0)
    # class App(tk.Tk):

Apps()

# def printData(conn: Connection):
#     table_type = 'BASE TABLE'
#     allddl: str = ''
#     cursor = conn.cursor()
#     cursor.execute('SELECT trim(upper(table_name)) as table_name FROM information_schema.tables')
#     # where table_type = ? ', '{table_type}')
#     result = cursor.fetchall()
#     for row in result:
#         #    print(f'{row.emp_code}: {row.emp_name}')
#         # print(f'{row.table_name}')
#         table_name = row.table_name
#         ddl = 'CREATE TABLE ' + table_name + '( \n'
#         cursor1 = conn.cursor()
#         result1 = cursor1.execute('SELECT upper(table_name) table_name,upper(column_name) column_name,'
#                                   'upper(data_type) data_type,upper(is_nullable) is_nullable,'
#                                   'character_maximum_length,numeric_precision,numeric_scale '
#                                   'FROM information_schema.columns '
#                                   'where trim(upper(table_name)) = ? order by ordinal_position', table_name)
#         for row1 in result1:
#             # if row1.column_name == 'VARCHAR':
#             #     data_type = row1.column_name + '( ' + row.character_maximum_length + ') '
#             # elif row1.column_name == 'INT':
#             #     data_type = row1.column_name
#             # elif row1.column_name == 'DECIMAL':
#             #     data_type = row1.column_name + '( ' + row.numeric_precision + ',' + row1.numeric_scale + ') '
#             # if row1.is_nullable[:] == 'NO':
#             #     isNull = ' NOT NULL'
#             # else:
#             #     isNull = ' '
#             #     ddl = (' ' + isNull + ', \n')
#             print(f'{row1.table_name}: {row1.column_name}: {row1.data_type}: {row1.is_nullable}:'
#                   '{row1.character_maximum_length}: {row1.numeric_precision}: {row1.numeric_scale}')
#         cursor1.close()
#
#     cursor.close()
